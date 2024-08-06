import csv
import openpyxl
import requests
import os
from openaigen import *
from fetchdata import * 
from S3export import *

# Retrieve environment variables
cluster = os.getenv('CLUSTER').lower()
tenant = os.getenv('TENANT')
bot = os.getenv('BOT_ID')
num_docs = int(os.getenv('NUM_DOCS'))
host_ip = "localhost"
date_time = os.getenv('TODAY')

print("--------------------")
print("Environment:", cluster)
print("Tenant:", tenant)
print("Source bot_id:", bot)
print("Date: ", date_time)

def iterate_docs(cluster, tenant, bot):
    csv_file_path = f'/logs/Info_cluster{cluster}_tenant{tenant}_botid{bot}.csv'

    # Check if the CSV file with the document keys exists
    if not os.path.isfile(csv_file_path):
        if not getDocKeys(tenant, bot):
            print("Incorrect combination of cluster/tenant/bot... Try again.")
            raise SystemExit
        
        print("First time accessing this bot, gathering all documents...")

    print("Iterating through the documents...")

    # Read the document keys and other information from the CSV file 
    with open(csv_file_path) as file_obj:
        reader_obj = csv.reader(file_obj) 
        row_count = 0 
        
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Process each document
        for row in reader_obj: 
            doc_key = int(row[0])
            source_url = row[1]
            title = row[2]
            section_url = f'http://{host_ip}:8088/tenant-server/v1/tenants/{tenant}/external-documents/retrieve-sections?documentKey={doc_key}&isCommitted=true'
            response = requests.get(section_url)

            #Checks if call was successful
            if response.status_code == 200:
                if title is None:
                    title = f'No Title. Document Key: {doc_key}'
                if source_url is None:
                    source_url = f'No URL. Document Key: {doc_key}'

                utterances = retrieve_data(tenant, doc_key, title, source_url, host_ip)
                
                # Write each utterance to the Excel sheet
                for utterance in utterances:
                    start_row = sheet.max_row + 1
                    for col_index, item in enumerate(utterance, start=1):
                        sheet.cell(row=start_row, column=col_index, value=item)
                        
                row_count += 1

            if row_count >= num_docs:
                break
            
        # Add headers to the Excel sheet
        sheet.cell(1, 1, "Doc Title")
        sheet.cell(1, 2, "Section Title")
        sheet.cell(1, 3, "URL")
        sheet.cell(1, 4, "Utterance/Phrase")
        
        # Save the Excel workbook
        workbook.save(f"../results/{cluster}_tenant{tenant}_botid{bot}_excel_{date_time}.xlsx")        
        
        # Upload the file to S3
        uploadFile_to_S3(cluster, tenant, bot)

#creates a csv file with information about the documents from the bot
def getDocKeys(tenant, botid):
    url = f"http://{host_ip}:8088/tenant-server/v1/tenants/{tenant}/external-documents/check-health?botId={botid}"
    try:
        response = json.loads(requests.get(url).text)
        print("Retrieved keys!!!")
    except Exception as e:
        print(f"Failed to retrieve keys: {e}")
        return False
    
    file_path = f"/logs/Info_cluster{cluster}_tenant{tenant}_botid{botid}.csv"
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Write the document keys and other information to the CSV file
    with open(file_path, 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile)
        for item in response:
            csvwriter.writerow([item.get("documentKey"), item.get("source"), item.get("title")])

    return True

print("...")
iterate_docs(cluster, tenant, bot)